# -*- coding: utf-8 -*-
#
# Copyright 2017  Inc
# @author ipqhjjybj 

'''
上传期货的数据
'''
import sys
import os
import six
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Font, Fill , Color
from openpyxl.styles import colors

sys.path.append("..")


from future import *

mubanDocPath = "D:/loopback/LoopBack/doc_muban/tuijin_analyse.xlsx"
symbols = ["ag","al","au","cu","ni","pb","sn","zn","bu","fg","l","ma","pp","ru","ta","v","a","c","cf","cs","jd","m","oi","p","rm","sr","y","hc","rb","i","j","jm","zc","if"]
GG_str = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR']
# 复制文件
def copyFile(dstPath,srcPath):
    f_in = open(srcPath , "rb")
    f_out = open(dstPath , "wb")
    f_out.write(f_in.read())
    f_in.close()
    f_out.close()

# 遍历获得所有 文件的具体地址
def getAllFilePath(srcPathDir):
    ret = []
    for parent,dirnames,filenames in os.walk(srcPathDir):
        for filename in filenames:              
            #print os.path.join(parent , filename)
            ret.append( os.path.join(parent , filename))
    return ret

# 通过部分字段，获得正确的文件路径
def getFilePath( arr , path_part):
    ret = ""
    for line in arr:
        if path_part in line:
            ret = line
            break
    return ret
# 通过部分字段，获得正确的文件路径
def getFilePath2( arr , path_part1 , path_part2):
    ret = ""
    for line in arr:
        if path_part1 in line and path_part2 in line:
            ret = line
            break
    return ret

# 获得平均收益
def getAverageRate( tuijin_path):
    line_num = 0
    cc_num = 0
    average_ybn_income_rate = 0.0
    average_ybw_income_rate = 0.0
    average_ybw_huiche = 0.0

    if os.path.exists(tuijin_path) == True:
        f = open(tuijin_path , "r")
        for line in f:
            line_num += 1
            if line_num >= 4:
                arr = line.strip().split(',')
                arr = [x.replace("%","") for x in arr]
                try:
                    average_ybn_income_rate += float(arr[4])
                    average_ybw_income_rate += float(arr[5])
                    average_ybw_huiche += float(arr[9])
                    cc_num += 1
                except Exception,ex:
                    print ex
        f.close()
        if cc_num > 0:
            average_ybn_income_rate = round(average_ybn_income_rate / cc_num / 100.0, 3)
            average_ybw_income_rate = round(average_ybw_income_rate / cc_num / 100.0, 3)
            average_ybw_huiche = round(average_ybw_huiche / cc_num / 100.0, 3)

    return average_ybn_income_rate , average_ybw_income_rate , average_ybw_huiche
    #return str(average_ybn_income_rate) + "%" , str(average_ybw_income_rate) + "%" , str(average_ybw_huiche) + "%" 

def makeAutoAnalyseReport( ybn_month = 12 , ybw_month = 3 ,dstPathName = "D:/loopback/LoopBack/doc_muban/a.xlsx" ,origin_report_path = "D:/loopback_period_dual_thrust_report_auto_produce/LoopBack/report/PeriodDualthrust"):
    global mubanDocPath , symbols , GG_str
    if os.path.exists(mubanDocPath) == True:
        copyFile( dstPathName , mubanDocPath)

    all_filenames = getAllFilePath(origin_report_path)

    wb = load_workbook(filename = dstPathName, keep_vba=False)

    index = 3
    for symbol in symbols:
        wt = wb["total"]
        filePath = getFilePath( all_filenames , "PushOnReport_based_on_%s_" % symbol)
        (average_ybn_income_rate , average_ybw_income_rate , average_ybw_huiche) = getAverageRate(filePath)

        wt["E"+str(index)] = ybn_month
        wt["F"+str(index)] = ybw_month
        wt["G"+str(index)] = average_ybn_income_rate
        wt["H"+str(index)] = average_ybw_income_rate
        wt["I"+str(index)] = average_ybw_huiche

        if abs(average_ybw_huiche) > 0.0:
            wt["K"+str(index)] = average_ybw_income_rate / abs(average_ybw_huiche)
        else:
            wt["K"+str(index)] = 0.0

        if average_ybw_income_rate > 0.05:
            wt["H"+str(index)].font = Font(color=colors.BLUE)
        if average_ybw_income_rate > 0.10:
            wt["H"+str(index)].font = Font(color=colors.RED)

        
        # 填充其他表格字段
        wt = wb[symbol]
        if os.path.exists(filePath) == True:
            line_num = 0
            aa_num = 2
            f = open(filePath , "r")
            for line in f:
                line_num += 1
                if line_num >= 4:
                    aa_num += 1
                    arr = line.strip().split(',')
                    for j in range(0, len(arr) ):
                        if '%' in arr[j]:
                            if '#J' in arr[j]:
                                wt[GG_str[j] + str(aa_num)] = 0.0
                            else:
                                s = arr[j].replace('%',"")
                                wt[GG_str[j] + str(aa_num)] = float(s) / 100.0
                        else:
                            if '/' in arr[j]:
                                wt[GG_str[j] + str(aa_num)] = str(arr[j])
                            else:
                                try:
                                    wt[GG_str[j] + str(aa_num)] = float(arr[j])
                                except Exception,ex:
                                    wt[GG_str[j] + str(aa_num)] = str(arr[j])
                        # wt[GG_str[j] + str(aa_num)].font = Font(color=colors.RED)
                        # wt[GG_str[j] + str(aa_num)].fill = PatternFill(fill_type=None,start_color=colors.BLUE,end_color=colors.BLUE)

            f.close()

        ybw_zdhc = 0.0

        filePath = getFilePath2( all_filenames , "%s-" % symbol , "push_backtesting_report.csv")
        if os.path.exists(filePath) == True:
            start_line_num = 50
            f = open(filePath , "r")
            for line in f:
                try:
                    line = line.strip().decode('utf-8')
                    line = (line + "\n")
                except Exception,ex:
                    line = "Label,净利润".decode('utf-8')

                arr = line.strip().split(',')
                if len(arr) > 1:
                    if "最大回撤".decode('utf-8') in arr[0] and '%' in arr[1]:
                        ybw_zdhc = float(arr[1].replace('%',"")) / 100.0
                    wt[GG_str[0] + str(start_line_num)] = arr[0]
                    wt[GG_str[1] + str(start_line_num)] = arr[1]
                start_line_num += 1
            f.close()

        filePath = getFilePath2( all_filenames , "%s-" % symbol , "best-figure-push-on.csv")
        if os.path.exists(filePath) == True:
            start_line_num = 50
            f = open(filePath , "r")
            for line in f:
                try:
                    line = line.strip().decode('utf-8')
                    line = (line + "\n")
                except Exception,ex:
                    line = "?,?\n"
                arr = line.strip().split(',')
                if len(arr) > 1:
                    wt[GG_str[6] + str(start_line_num)] = arr[0]
                    #wt[GG_str[7] + str(start_line_num)] = str(arr[1])
                    try:
                        wt[GG_str[7] + str(start_line_num)] = float(arr[1])
                    except Exception,ex:
                        wt[GG_str[7] + str(start_line_num)] = arr[1]
                start_line_num += 1
            f.close()

        wt = wb["total"]
        wt["J"+str(index)] = ybw_zdhc
        if abs(ybw_zdhc) > 0.0:
            wt["L"+str(index)] = average_ybw_income_rate  / abs(ybw_zdhc)
        else:
            wt["L"+str(index)] = 0.0
        index += 1

    wb.save(dstPathName)

def main():
    #makeAutoAnalyseReport(ybn_month = 12 , ybw_month = 3  , dstPathName = "D:/loopback/LoopBack/doc_muban/periodDualThrust_12_3.xlsx", origin_report_path = "D:/loopback_period_dual_thrust_12_3/LoopBack/report/PeriodDualthrust")
    #makeAutoAnalyseReport(ybn_month = 24 , ybw_month = 3  , dstPathName = "D:/loopback/LoopBack/doc_muban/periodDualThrust_24_3.xlsx", origin_report_path = "D:/loopback_period_dual_thrust_24_3/LoopBack/report/PeriodDualthrust")
    #makeAutoAnalyseReport(ybn_month = 12 , ybw_month = 3  , dstPathName = "D:/loopback/LoopBack/doc_muban/Sclw_jdzs_12_3.xlsx", origin_report_path = "D:/loopback_period_sclw_jdzs_12_3/LoopBack/report/Sclw_JdzsStrategy")
    #makeAutoAnalyseReport(ybn_month = 24 , ybw_month = 3  , dstPathName = "D:/loopback/LoopBack/doc_muban/Sclw_jdzs_24_3.xlsx", origin_report_path = "D:/loopback_period_sclw_jdzs_24_3/LoopBack/report/Sclw_JdzsStrategy")
    makeAutoAnalyseReport(ybn_month = 24 , ybw_month = 6  , dstPathName = "D:/loopback/LoopBack/doc_muban/Sclw_jdzs_24_6.xlsx", origin_report_path = "D:/loopback_period_sclw_jdzs_24_6/LoopBack/report/Sclw_JdzsStrategy")
    
    
if __name__ == '__main__':
    main()